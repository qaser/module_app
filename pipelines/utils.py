import os
import re
import sys
from datetime import date, datetime
from pathlib import Path

import django
from openpyxl import load_workbook

BASE_URL = r"G:\WorkDocuments\Dev\module_app"
# BASE_URL = r"D:\Development\module_a5pp"

# Настройка Django окружения
sys.path.append(BASE_URL)
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "module_app.settings")
django.setup()


from pipelines.models import (
    Anomaly, Bend, Diagnostics, Pipe, Tube, TubeUnit, TubeVersion)

Bend.objects.all().delete()
Tube.objects.all().delete()
TubeVersion.objects.all().delete()
TubeUnit.objects.all().delete()
Anomaly.objects.all().delete()
Diagnostics.objects.all().delete()

''' импорт труб
0 tube_num
1 odometr_data
2 tube_length
3 tube_type
4 thickness
5 weld_position
6 yield_strength
7 tear_strength
8 category
9 reliability_material
10 comment
11 steel_grade
'''

''' импорт элементов обустройства
0 odometr_data
1 tube_num
3 unit_type
4 description
5 comment
'''

''' импорт аномалий
0 odometr_data
1 from_left_weld_to_max
2 from_left_weld_to_start
3 from_right_weld_to_max
4 from_right_weld_to_start
5 from_long_weld_to_start
6 from_long_weld_to_max
7 from_long_weld_to_center
8 min_distance_to_long_weld
9 min_distance_to_circ_weld
10 tube_num
11 anomaly_description
12 anomaly_nature
13 size_class
14 start_point_orientation
15 max_point_orientation
16 center_orientation
17 anomaly_length
18 anomaly_width
19 anomaly_depth
20 location
21 comment
22 safe_pressure_coefficient
23 danger_level
'''

''' импорт элементов отводов
0 start_point
1 end_point
2 tube_num
3 segment_count
4 radius
5 bend_angle
6 projection_angle
7 bend_type
8 direction
9 comment
10 latitude
11 longitude
12 altitude
13 safety_status
'''

HEADER_KEYWORDS = ["Номер трубы", "Толщина", "Тип трубы", 'Расстояние, м']
DATA = [
    {
        'name': 'nord_uu',
        'report_type': 'new',
        'pipe_ranges': {
            2: "2932 - 5108",  # номер участка и диапазон труб
            3: "5110 - 7351",
            4: "7353 - 7466",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\nord_uu\tubes_nord_uu.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\nord_uu\tubeunits_nord_uu.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\nord_uu\anomalies_nord_uu.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\nord_uu\bends_nord_uu.xlsx",
        'diagnostics_start': '04.08.2025',
        'diagnostics_end': '07.08.2025',
    },
    {
        'name': 'south_uu',
        'report_type': 'new',
        'pipe_ranges': {
            5: "2 - 98б",  # номер участка и диапазон труб
            6: "133а - 2732",
            7: "2734 - 5452",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\south_uu\tubes_south_uu.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\south_uu\tubeunits_south_uu.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\south_uu\anomalies_south_uu.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\south_uu\bends_south_uu.xlsx",
        'diagnostics_start': '24.06.2023',
        'diagnostics_end': '29.06.2023',
    },
    {
        'name': 'nord_c1',
        'report_type': 'new',
        'pipe_ranges': {
            8: "2987 - 5157",  # номер участка и диапазон труб
            9: "5159 - 7421",
            10: "7423 - 7498б",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\nord_c1\tubes_nord_c1.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\nord_c1\tubeunits_nord_c1.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\nord_c1\anomalies_nord_c1.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\nord_c1\bends_nord_c1.xlsx",
        'diagnostics_start': '06.03.2024',
        'diagnostics_end': '24.03.2024',
    },
    {
        'name': 'south_uc1',
        'report_type': 'new',
        'pipe_ranges': {
            17: "2 - 117а",  # номер участка и диапазон труб
            18: "118б - 2803",
            19: "2805 - 5600",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\south_uc1\tubes_south_uc1.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\south_uc1\tubeunits_south_uc1.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\south_uc1\anomalies_south_uc1.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\south_uc1\bends_south_uc1.xlsx",
        'diagnostics_start': '11.08.2025',
        'diagnostics_end': '15.08.2025',
    },
    {
        'name': 'nord_c2',
        'report_type': 'new',
        'pipe_ranges': {
            14: "2942а - 5194а",  # номер участка и диапазон труб
            15: "5196 - 7479",
            16: "7481 - 7538а",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\nord_c2\tubes_nord_c2.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\nord_c2\tubeunits_nord_c2.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\nord_c2\anomalies_nord_c2.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\nord_c2\bends_nord_c2.xlsx",
        'diagnostics_start': '04.04.2025',
        'diagnostics_end': '07.04.2025',
    },
    {
        'name': 'south_c2',
        'report_type': 'new',
        'pipe_ranges': {
            14: "2 - 131а",  # номер участка и диапазон труб
            15: "133а - 2828а",
            16: "2830 - 5596",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\south_c2\tubes_south_c2.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\south_c2\tubeunits_south_c2.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\south_c2\anomalies_south_c2.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\south_c2\bends_south_c2.xlsx",
        'diagnostics_start': '19.08.2024',
        'diagnostics_end': '23.08.2024',
    },
    {
        'name': 'nord_e1',
        'report_type': 'new',
        'pipe_ranges': {
            20: "2917 - 5172",  # номер участка и диапазон труб
            21: "5174 - 7432",
            22: "7434 - 7534",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\nord_e1\tubes_nord_e1.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\nord_e1\tubeunits_nord_e1.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\nord_e1\anomalies_nord_e1.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\nord_e1\bends_nord_e1.xlsx",
        'diagnostics_start': '13.02.2024',
        'diagnostics_end': '17.02.2024',
    },
    {
        'name': 'south_e1',
        'report_type': 'new',
        'pipe_ranges': {
            23: "2 - 102",  # номер участка и диапазон труб
            24: "104 - 2809",
            25: "2811 - 5575",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\south_e1\tubes_south_e1.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\south_e1\tubeunits_south_e1.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\south_e1\anomalies_south_e1.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\south_e1\bends_south_e1.xlsx",
        'diagnostics_start': '04.12.2024',
        'diagnostics_end': '10.12.2024',
    },
    # {
    #     'name': 'nord_e2',
    #     'report_type': 'new',
    #     'pipe_ranges': {
    #         26: "2936 - 5240б",  # номер участка и диапазон труб
    #         27: "5242а - 7496",
    #         28: "7498 - 7651",
    #     },
    #     'filepath_tubes': BASE_URL + r"\fixtures\data\nord_e2\tubes_nord_e2.xlsx",
    #     'filepath_units': BASE_URL + r"\fixtures\data\nord_e2\tubeunits_nord_e2.xlsx",
    #     'filepath_anomalies': BASE_URL + r"\fixtures\data\nord_e2\anomalies_nord_e2.xlsx",
    #     'filepath_bends': BASE_URL + r"\fixtures\data\nord_e2\bends_nord_e2.xlsx",
    #     'diagnostics_start': '13.08.2024',
    #     'diagnostics_end': '17.08.2024',
    # },
    {
        'name': 'south_e2',
        'report_type': 'new',
        'pipe_ranges': {
            29: "2 - 54",  # номер участка и диапазон труб
            30: "56 - 2795",
            31: "2797 - 5581",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\south_e2\tubes_south_e2.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\south_e2\tubeunits_south_e2.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\south_e2\anomalies_south_e2.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\south_e2\bends_south_e2.xlsx",
        'diagnostics_start': '17.01.2023',
        'diagnostics_end': '21.01.2023',
    },
    {
        'name': 'nord_zg',
        'report_type': 'new',
        'pipe_ranges': {
            32: "2936 - 5240б",  # номер участка и диапазон труб
            33: "5242а - 7496",
            34: "7498 - 7651",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\nord_zg\tubes_nord_zg.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\nord_zg\tubeunits_nord_zg.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\nord_zg\anomalies_nord_zg.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\nord_zg\bends_nord_zg.xlsx",
        'diagnostics_start': '14.12.2023',
        'diagnostics_end': '18.12.2023',
    },
    {
        'name': 'south_zg',
        'report_type': 'new',
        'pipe_ranges': {
            35: "2 - 32",  # номер участка и диапазон труб
            36: "34 - 2753",
            37: "2755 - 5545",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\south_zg\tubes_south_zg.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\south_zg\tubeunits_south_zg.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\south_zg\anomalies_south_zg.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\south_zg\bends_south_zg.xlsx",
        'diagnostics_start': '05.10.2023',
        'diagnostics_end': '09.10.2023',
    },
    {
        'name': 'nord_t1',
        'report_type': 'new',
        'pipe_ranges': {
            38: "2947 - 5257",  # номер участка и диапазон труб
            39: "5259 - 7430",
            40: "7432 - 7535",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\nord_t1\tubes_nord_t1.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\nord_t1\tubeunits_nord_t1.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\nord_t1\anomalies_nord_t1.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\nord_t1\bends_nord_t1.xlsx",
        'diagnostics_start': '23.08.2023',
        'diagnostics_end': '28.08.2023',
    },
    {
        'name': 'south_yat1',
        'report_type': 'new',
        'pipe_ranges': {
            41: "2 - 138",  # номер участка и диапазон труб
            42: "140 - 2924",
            43: "2926 - 5718",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\south_yat1\tubes_south_yat1.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\south_yat1\tubeunits_south_yat1.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\south_yat1\anomalies_south_yat1.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\south_yat1\bends_south_yat1.xlsx",
        'diagnostics_start': '11.10.2024',
        'diagnostics_end': '15.10.2024',
    },
    {
        'name': 'nord_yat2',
        'report_type': 'new',
        'pipe_ranges': {
            44: "2909 - 5217",  # номер участка и диапазон труб
            45: "5219 - 7417",
            46: "7419 - 7503",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\nord_yat2\tubes_nord_yat2.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\nord_yat2\tubeunits_nord_yat2.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\nord_yat2\anomalies_nord_yat2.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\nord_yat2\bends_nord_yat2.xlsx",
        'diagnostics_start': '04.02.2025',
        'diagnostics_end': '07.02.2025',
    },
    {
        'name': 'south_yat2',
        'report_type': 'new',
        'pipe_ranges': {
            47: "2аб - 172",  # номер участка и диапазон труб
            48: "174 - 2910",
            49: "2912 - 5756",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\south_yat2\tubes_south_yat2.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\south_yat2\tubeunits_south_yat2.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\south_yat2\anomalies_south_yat2.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\south_yat2\bends_south_yat2.xlsx",
        'diagnostics_start': '16.10.2025',
        'diagnostics_end': '18.10.2025',
    },
    {
        'name': 'nord_pov',
        'report_type': 'new',
        'pipe_ranges': {
            50: "3346 - 5511",  # номер участка и диапазон труб
            51: "5513 - 7791",
            52: "7793 - 7943",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\nord_pov\tubes_nord_pov.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\nord_pov\tubeunits_nord_pov.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\nord_pov\anomalies_nord_pov.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\nord_pov\bends_nord_pov.xlsx",
        'diagnostics_start': '01.07.2023',
        'diagnostics_end': '05.07.2023',
    },
    {
        'name': 'south_pov',
        'report_type': 'new',
        'pipe_ranges': {
            53: "4 - 135",  # номер участка и диапазон труб
            54: "137а - 2873а",
            55: "2875а - 5742",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\south_pov\tubes_south_pov.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\south_pov\tubeunits_south_pov.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\south_pov\anomalies_south_pov.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\south_pov\bends_south_pov.xlsx",
        'diagnostics_start': '20.02.2024',
        'diagnostics_end': '26.02.2024',
    },
    {
        'name': 'nord_srtou',
        'report_type': 'new',
        'pipe_ranges': {
            56: "3345 - 5526",  # номер участка и диапазон труб
            57: "5528 - 7825",
            58: "7827 - 7988",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\nord_srtou\tubes_nord_srtou.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\nord_srtou\tubeunits_nord_srtou.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\nord_srtou\anomalies_nord_srtou.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\nord_srtou\bends_nord_srtou.xlsx",
        'diagnostics_start': '07.02.2023',
        'diagnostics_end': '11.02.2023',
    },
    {
        'name': 'south_srtou',
        'report_type': 'new',
        'pipe_ranges': {
            59: "2 - 130",  # номер участка и диапазон труб
            60: "132 - 2874",
            61: "2876 - 5662",
        },
        'filepath_tubes': BASE_URL + r"\fixtures\data\south_srtou\tubes_south_srtou.xlsx",
        'filepath_units': BASE_URL + r"\fixtures\data\south_srtou\tubeunits_south_srtou.xlsx",
        'filepath_anomalies': BASE_URL + r"\fixtures\data\south_srtou\anomalies_south_srtou.xlsx",
        'filepath_bends': BASE_URL + r"\fixtures\data\south_srtou\bends_south_srtou.xlsx",
        'diagnostics_start': '09.02.2025',
        'diagnostics_end': '12.02.2025',
    },
]


def is_header(row):
    """Определяем, является ли строка заголовком таблицы."""
    row_str = " ".join([str(x) for x in row if x])
    return any(k in row_str for k in HEADER_KEYWORDS)


def clean_cell_value(value):
    """Очищает значение от вертикальных линий и конвертирует в float"""
    if isinstance(value, str):
        # Убираем вертикальные линии с начала и конца строки
        value = value.strip('|')
        # Заменяем запятую на точку для десятичных чисел
        value = value.replace(',', '.')
        # Убираем пробелы
        value = value.strip()

    try:
        return float(value) if value else None
    except (ValueError, TypeError):
        return None


def extract_number_and_suffix(tube_num: str):
    """Извлекаем числовую часть и суффикс (например, 2941а → 2941, 'а')."""
    if not tube_num:
        return None, ""
    m = re.match(r"(\d+)(.*)", str(tube_num).strip())
    if not m:
        return None, ""
    return int(m.group(1)), m.group(2).strip().lower()


def parse_range(range_str: str):
    """Парсим диапазон '2941а - 5195' → (2941, 5195)."""
    parts = [p.strip() for p in range_str.split("-")]
    start, _ = extract_number_and_suffix(parts[0])
    end, _ = extract_number_and_suffix(parts[1])
    return start, end


def find_pipe_for_tube(tube_num, pipe_ranges):
    """Определяем участок (Pipe) по номеру трубы."""
    num, _ = extract_number_and_suffix(tube_num)
    if num is None:
        return None
    for pipe_id, range_str in pipe_ranges.items():
        start, end = parse_range(range_str)
        # диапазон закрытый (включаем края)
        if start <= num <= end:
            try:
                return Pipe.objects.get(id=pipe_id)
            except Pipe.DoesNotExist:
                print(f"⚠️ Pipe id={pipe_id} не найден")
                return None
    return None


def get_or_create_diagnostics_for_pipes(pipe_ids, start_str, end_str):
    """Создаёт (или получает) один объект Diagnostics, связанный со всеми участками."""
    start_date = datetime.strptime(start_str, "%d.%m.%Y").date()
    end_date = datetime.strptime(end_str, "%d.%m.%Y").date()

    # Пробуем найти существующую диагностику
    diagnostics = Diagnostics.objects.filter(
        start_date=start_date,
        end_date=end_date,
    ).first()

    if not diagnostics:
        # Создаём без вызова full_clean()
        diagnostics = Diagnostics.objects.create(
            start_date=start_date,
            end_date=end_date,
            description=f"Диагностика участков {', '.join(map(str, pipe_ids))} ({start_date}–{end_date})"
        )
        print(f"🧾 Создан новый объект диагностики (id={diagnostics.id})")
    else:
        print(f"ℹ️ Найдена существующая диагностика ({start_date}–{end_date}) id={diagnostics.id}")

    # Теперь безопасно добавляем участки (M2M)
    for pipe_id in pipe_ids:
        try:
            pipe = Pipe.objects.get(id=pipe_id)
            diagnostics.pipes.add(pipe)
        except Pipe.DoesNotExist:
            print(f"⚠️ Участок id={pipe_id} не найден")

    # Теперь вызываем clean(), чтобы модель прошла проверку
    diagnostics.full_clean()
    diagnostics.save()
    return diagnostics


def import_tubes(
        filepath,
        filepath_units,
        filepath_anomalies,
        filepath_bends,
        pipe_ranges: dict,
        diagnostics_start: str,
        diagnostics_end: str
    ):
    TUBE_TYPE_MAP = {
        '1Ш': 'one',
        '2Ш': 'two',
        'СШ': 'spiral',
        'БШ': 'without',
    }
    print(f"📘 Импорт из файла: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    # создаём/находим диагностику, связанную со всеми участками
    diagnostics = get_or_create_diagnostics_for_pipes(pipe_ranges.keys(), diagnostics_start, diagnostics_end)
    header_found = False
    created_versions = 0
    created_tubes = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if is_header(row):
            header_found = True
            continue
        if not header_found:
            continue
        if not row[1]:  # пустая строка
            continue
        if is_header(row):  # повторная шапка
            continue

        tube_num = str(row[0]).strip()
        pipe = find_pipe_for_tube(tube_num, pipe_ranges)
        if pipe is None:
            continue

        # --- Tube ---
        tube, created = Tube.objects.get_or_create(
            pipe=pipe,
            tube_num=tube_num,
            defaults={"active": True, "installed_date": date.today()},
        )
        if created:
            created_tubes += 1

        # --- TubeVersion ---
        raw_type = str(row[3]).strip() if row[3] else None
        tube_type = TUBE_TYPE_MAP.get(raw_type, 'without')
        if tube_type is None:
            raise ValueError(f'Неизвестный тип трубы: {raw_type}')
        try:
            TubeVersion.objects.create(
                tube=tube,
                diagnostics=diagnostics,
                version_type="diagnostic",
                date=diagnostics.end_date,
                tube_num=tube_num,
                odometr_data = float(row[1]) if row[1] else 0,
                tube_length=float(row[2]) if row[2] else 0,
                tube_type=tube_type,
                thickness=float(row[4]) if row[4] else 0,
                weld_position=str(row[5]).strip() if row[5] else None,
                yield_strength=float(row[6]) if row[6] else 0,
                tear_strength=float(row[7]) if row[7] else 0,
                category=str(row[8]).strip() if row[8] else "II",
                reliability_material=float(row[9]) if row[9] else None,
                comment=str(row[10]).strip() if row[10] else None,
                steel_grade=str(row[11]).strip() if row[11] else None,
                # reliability_coef=float(row[12]) if row[12] else None,
                # reliability_pressure=float(row[13]) if row[13] else None,
                # working_conditions=float(row[14]) if row[14] else None,
                # impact_strength=float(row[15]) if row[15] else None,
                # from_reference_start=str(row[15]).strip() if row[15] else None,
                # to_reference_end=str(row[16]).strip() if row[16] else None,
            )
            created_versions += 1
        except Exception as e:
            print(f"⚠️ Ошибка в строке {i}: {e}")
            continue
    print(f"\n✅ Импорт завершён:")
    print(f"  • труб создано — {created_tubes}")
    print(f"  • версий создано — {created_versions}")
    print(f"  • диагностика ID={diagnostics.id}, диапазон {diagnostics_start}–{diagnostics_end}")
    import_tube_units(filepath_units, diagnostics)
    import_anomalies(filepath_anomalies, diagnostics)
    import_bends(filepath_bends, diagnostics)


def import_tube_units(filepath, diagnostics):
    """
    Импорт элементов обустройства и привязка к TubeVersion последней диагностики.
    """
    print(f"📘 Импорт элементов обустройства из: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    header_found = False
    created_units = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Пропуск мусора и шапок
        if not any(row):
            continue
        row_str = " ".join([str(x) for x in row if x])
        if any(keyword in row_str.lower() for keyword in ["тип", "одометр", "трубы"]):
            header_found = True
            continue
        if not header_found:
            continue
        try:
            tube_num = str(row[1]).strip()
            unit_type_raw = str(row[3]).strip().lower() if row[3] else None
            odometr = float(row[0]) if row[0] else None
            description = str(row[4]).strip() if row[4] else None
            comment = str(row[5]).strip() if row[5] else None
        except Exception as e:
            print(f"⚠️ Ошибка парсинга строки {i}: {e}")
            continue

        tube_version = TubeVersion.objects.filter(diagnostics=diagnostics, tube_num=tube_num).first()
        if not tube_version:
            # print(f"⚠️ Версия для трубы {tube_num} и диагностики {diagnostics.id} не найдена")
            continue

        # определяем тип элемента
        UNIT_TYPE_MAP = {
            "кран": "valv",
            "отвод": "offt",
            "врезка": "offt",
            "тройник": "tee",
            "эхз": "cpco",
            "окно": "wiwd",
            "футляр-начало": "casb",
            "футляр-конец": "case",
            "маркер": "mark",
            "пригруз": "anch",
            "обустройство": "pfix",
        }
        unit_type = "pfix"  # по умолчанию
        for k, v in UNIT_TYPE_MAP.items():
            if v in unit_type_raw:
                unit_type = v
                break
        # создаём элемент
        TubeUnit.objects.create(
            tube=tube_version,
            unit_type=unit_type,
            odometr_data=odometr,
            description=description,
            comment=comment,
        )
        created_units += 1

    print(f"\n✅ Импорт элементов завершён:")
    print(f"  • создано элементов: {created_units}")
    print(f"  • диагностика: {diagnostics.id} ({diagnostics.start_date} — {diagnostics.end_date})")


def import_anomalies(filepath, diagnostics):
    ANOMALY_NATURE_MAP = {
        'Аномалия кольцевого шва': 'gwan',
        'Аномалия спирального шва': 'swan',
        'Аномалия продольного шва': 'lwan',
        'Механическое повреждение': 'goug',
        'Коррозия': 'corr',
        'Вмятина': 'dent',
        'Гофр': 'wrin',
        'Технологический дефект': 'artd',
        'Заводской дефект': 'mian',
        'Зона продольных трещин': 'scc',
        'Трещина на продольном шве': 'lwcr',
        'Эллипсность': 'oval',
    }
    SIZE_CLASS_MAP = {
        'Не указан': '',
        'Обширный': 'gene',
        'Каверна': 'pitt',
        'Поперечная канавка': 'cigr',
        'Продольная канавка': 'axgr',
        'Продольный паз': 'axsl',
        'Поперечный паз': 'cisl',
    }
    LOCATION_MAP = {
        'INT': 'int',
        'EXT': 'ext',
        'MID': 'mid',
        'N/A': 'n/a',
    }
    print(f"📘 Импорт аномалий из: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    header_found = False
    created_anomalies = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # пропускаем пустые строки
        if not any(row):
            continue

        # определяем строку с заголовками
        row_str = " ".join([str(x) for x in row if x])
        if "расстояние" in row_str.lower() and "шва" in row_str.lower():
            header_found = True
            continue
        if not header_found:
            continue

        try:
            # индекс колонок по структуре файла
            odometr_data = float(row[0]) if row[0] else None
            from_left_weld_to_max = float(row[1]) if row[1] else None
            from_left_weld_to_start = float(row[2]) if row[2] else None
            from_right_weld_to_max = float(row[3]) if row[3] else None
            from_right_weld_to_start = float(row[4]) if row[4] else None
            from_long_weld_to_start = float(row[5]) if row[5] else None
            from_long_weld_to_max = float(row[6]) if row[6] else None
            from_long_weld_to_center = float(row[7]) if row[7] else None
            min_distance_to_long_weld = float(row[8]) if row[8] else None
            min_distance_to_circ_weld = float(row[9]) if row[9] else None
            tube_num = str(row[10]).strip() if row[10] else None
            anomaly_description = str(row[11]) if row[11] else None
            anomaly_nature = str(row[12]).strip().lower() if row[12] else None  # "Характер особенности"
            size_class = str(row[13]).strip().lower() if row[13] else None      # "Класс размера"
            start_point_orientation = str(row[14]) if row[14] else None
            max_point_orientation = str(row[15]) if row[15] else None
            center_orientation = str(row[16]) if row[16] else None
            anomaly_length = int(row[17]) if row[17] else None
            anomaly_width = int(row[18]) if row[18] else None
            anomaly_depth = float(row[19]) if row[19] else None
            location = str(row[20]).strip().lower() if row[20] else None        # "Расположение"
            comment = str(row[21]) if row[21] else None
            safe_pressure_coefficient = float(row[22]) if row[22] else None
            danger_level = str(row[23]).strip() if len(row) > 23 and row[23] else None

        except Exception as e:
            print(f"⚠️ Ошибка парсинга строки {i}: {e}")
            continue

        if not tube_num:
            continue

        version = TubeVersion.objects.filter(tube_num=tube_num, diagnostics=diagnostics).first()
        if not version:
            continue

        # создаём аномалию
        try:
            Anomaly.objects.create(
                tube=version,
                odometr_data=odometr_data,
                from_left_weld_to_max=from_left_weld_to_max,
                from_left_weld_to_start=from_left_weld_to_start,
                from_right_weld_to_max=from_right_weld_to_max,
                from_right_weld_to_start=from_right_weld_to_start,
                from_long_weld_to_start=from_long_weld_to_start,
                from_long_weld_to_max=from_long_weld_to_max,
                from_long_weld_to_center=from_long_weld_to_center,
                min_distance_to_long_weld=min_distance_to_long_weld,
                min_distance_to_circ_weld=min_distance_to_circ_weld,
                start_point_orientation=start_point_orientation,
                max_point_orientation=max_point_orientation,
                center_orientation=center_orientation,
                anomaly_nature=anomaly_nature,
                anomaly_description=anomaly_description,
                size_class=size_class,
                location=location,
                anomaly_length=anomaly_length,
                anomaly_width=anomaly_width,
                anomaly_depth=anomaly_depth,
                comment=comment,
                safe_pressure_coefficient=safe_pressure_coefficient,
                danger_level=danger_level,
            )
            created_anomalies += 1
        except Exception as e:
            print(f"⚠️ Ошибка создания аномалии (строка {i}): {e}")

    print(f"\n✅ Импорт аномалий завершён:")
    print(f"  • создано записей: {created_anomalies}")
    print(f"  • диагностика: {diagnostics.id} ({diagnostics.start_date} — {diagnostics.end_date})")


def import_bends(filepath, diagnostics):
    # Словари для маппинга текстовых значений на ключи choices
    BEND_TYPE_MAP = {
        'Упруго-пластический изгиб': 'elastic_plastic',
        'Отвод холодного гнутья': 'cold_bend',
        'Отвод сегментный': 'segment_bend',
    }
    DIRECTION_MAP = {
        'Вертикальная': 'vertical',
        'Горизонтальная': 'horizontal',
    }
    print(f"📘 Импорт отводов из: {filepath}")
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    header_found = False
    created_bends = 0

    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # пропускаем пустые строки
        if not any(row):
            continue

        # определяем строку с заголовками
        row_str = " ".join([str(x) for x in row if x])
        if "начало" in row_str.lower() and "конец" in row_str.lower() and "номер трубы" in row_str.lower():
            header_found = True
            continue
        if not header_found:
            continue

        try:
            # индекс колонок по структуре файла
            start_point = float(row[0]) if row[0] else None
            end_point = float(row[1]) if row[1] else None
            tube_num = str(row[2]).strip() if row[2] else None
            segment_count = int(row[3]) if row[3] else None
            radius = float(row[4]) if row[4] else None
            bend_angle = float(row[5]) if row[5] else None
            projection_angle = str(row[6]) if row[6] else None

            # Поля choices - маппинг текста на ключи
            bend_type_text = str(row[7]).strip() if row[7] else None  # "Тип"
            direction_text = str(row[8]).strip() if row[8] else None  # "Направление"
            # Преобразуем текстовые значения в ключи choices
            bend_type = BEND_TYPE_MAP.get(bend_type_text)
            direction = DIRECTION_MAP.get(direction_text)

            comment = str(row[9]) if row[9] else None

            # Географические координаты (если есть в данных)
            latitude = float(row[10]) if len(row) > 10 and row[10] else None
            longitude = float(row[11]) if len(row) > 11 and row[11] else None
            altitude = float(row[12]) if len(row) > 12 and row[12] else None

            # Статус безопасности
            safety_status = str(row[13]) if len(row) > 13 and row[13] else None

        except Exception as e:
            print(f"⚠️ Ошибка парсинга строки {i}: {e}")
            continue

        if not tube_num or not start_point or not end_point:
            print(f"⚠️ Пропуск строки {i}: отсутствуют обязательные данные (труба, начало, конец)")
            continue

        version = TubeVersion.objects.filter(tube_num=tube_num, diagnostics=diagnostics).first()
        if not version:
            continue
        # создаём отвод
        try:
            bend = Bend.objects.create(
                tube=version,
                start_point=start_point,
                end_point=end_point,
                tube_num=tube_num,
                segment_count=segment_count,
                radius=radius,
                bend_angle=bend_angle,
                projection_angle=projection_angle,
                bend_type=bend_type,
                direction=direction,
                comment=comment,
                latitude=latitude,
                longitude=longitude,
                altitude=altitude,
                safety_status=safety_status,
            )
            # Вызываем save для автоматического вычисления radius_in_diameters и парсинга комментария
            bend.save()
            created_bends += 1
            # Отладочная информация для первых нескольких записей
            if created_bends <= 5:
                print(f"✅ Создан отвод: труба {tube_num}, тип: {bend_type_text} -> {bend_type}, направление: {direction_text} -> {direction}")

        except Exception as e:
            print(f"⚠️ Ошибка создания отвода (строка {i}): {e}")
            print(f"   Параметры: type={bend_type}, direction={direction}")
    print(f"\n✅ Импорт отводов завершён:")
    print(f"  • создано записей: {created_bends}")
    print(f"  • диагностика: {diagnostics.id} ({diagnostics.start_date} — {diagnostics.end_date})")


if __name__ == "__main__":
    for diagnostic_data in DATA:
        pipe_ranges = diagnostic_data.get('pipe_ranges')
        diagnostics_start = diagnostic_data.get('diagnostics_start')
        diagnostics_end = diagnostic_data.get('diagnostics_end')
        filepath_tubes = diagnostic_data.get('filepath_tubes')
        filepath_units = diagnostic_data.get('filepath_units')
        filepath_anomalies = diagnostic_data.get('filepath_anomalies')
        filepath_bends = diagnostic_data.get('filepath_bends')
        report_type = diagnostic_data.get('report_type')
        import_tubes(
            filepath_tubes,
            filepath_units,
            filepath_anomalies,
            filepath_bends,
            pipe_ranges,
            diagnostics_start,
            diagnostics_end
        )
